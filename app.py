from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from models import db, Service, User
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Border, Side
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect

app = Flask(__name__)

import secrets

def generate_secret_key(length=24):
    """
    Генерирует безопасный SECRET_KEY для Flask-приложения.

    :param length: Длина ключа (по умолчанию 24 символа, подходит для Flask)
    :return: Строка, представляющая случайный SECRET_KEY
    """
    return secrets.token_hex(length)

app.config['SECRET_KEY'] = generate_secret_key()

import os
app.secret_key = os.urandom(24)  # Генерирует случайный ключ длиной 24 байта
csrf_token = CSRFProtect(app)

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:enigma1418@172.18.11.104/mdtomskbot'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Изменим сообщение о необходимости входа
login_manager.login_message = 'Для доступа к этой странице, пожалуйста, войдите в систему.'
login_manager.login_message_category = 'info'  # Можно задать категорию, например, 'info', 'warning', 'danger'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get_or_404(int(user_id))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        new_user = User(username=username)
        new_user.set_password(password)  # Предполагаем, что метод set_password уже реализован
        db.session.add(new_user)
        db.session.commit()
        flash('Registration successful!', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, HiddenField
from wtforms.validators import DataRequired

class LoginForm(FlaskForm):
    csrf_token = HiddenField()
    submit = SubmitField('Submit')  # Кнопка отправки обязательно должна быть!

mas = {}

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Вход успешен!', 'success')
            if not current_user.username in mas:
                mas[current_user.username] = form
                # csrf_token = form
            # else:
            #     print('PFDSFDSFDSFSDFDSFDS')
            #     csrf_token = mas[current_user.username]
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
            # # print('POPAL_N', csrf_token)
            # return render_template('index.html', form=csrf_token)
        flash('Неверное имя пользователя или пароль!', 'danger')
    return render_template('login.html', form=form)

def skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token):
    per_page = 20

    import re

    # Регулярное выражение для формата "DD.MM.YYYY"
    pattern_dd_mm_yyyy = r'\b\d{2}\.\d{2}\.\d{4}\b'

    # Регулярное выражение для формата "YYYY-MM-DD"
    pattern_yyyy_mm_dd = r'\b\d{4}-\d{2}-\d{2}\b'

    # Сначала выбираем все уникальные значения year из базы
    all_years = db.session.query(Service.year,).distinct().all()
    all_years_date_number_no_one = db.session.query(Service.date_number_no_one).distinct().all()

    """----------------------------------"""

    service_years = []
    empty_found = False  # Флаг для отслеживания пустых строк

    for year_value in all_years:
        year_str = year_value[0]  # Извлекаем само значение year

        # Проверяем на пустую строку
        if not year_str:
            empty_found = True

        # Поиск всех дат в строке
        match_dd_mm_yyyy = re.findall(pattern_dd_mm_yyyy, year_str)
        match_yyyy_mm_dd = re.findall(pattern_yyyy_mm_dd, year_str)

        # Извлечение годов из найденных дат
        service_years.extend([date_str[-4:] for date_str in match_dd_mm_yyyy])  # Годы из формата "DD.MM.YYYY"
        service_years.extend([date_str[:4] for date_str in match_yyyy_mm_dd])    # Годы из формата "YYYY-MM-DD"

    # Оставляем только уникальные годы и преобразуем в целые числа
    service_years = list(set(int(year) for year in service_years if year.isdigit()))

    # Сортируем годы в порядке возрастания
    service_years.sort()

    # Если были пустые строки, добавляем None
    if empty_found:
        service_years.insert(0, None)

    # Преобразуем все значения в строки
    service_years = [str(year) for year in service_years]

    # service_years = db.session.query(db.func.year(Service.year)).distinct().all()
    # service_years = [str(year[0]) for year in service_years]

    """----------------------------------"""

    """----------------------------------"""

    service_date_number_no_one = []
    empty_found_date_number_no_one = False  # Флаг для отслеживания пустых строк

    for year_value in all_years_date_number_no_one:
        year_str = year_value[0]  # Извлекаем само значение year

        # Проверяем на пустую строку
        if not year_str:
            empty_found_date_number_no_one = True

        # Поиск всех дат в строке
        match_dd_mm_yyyy = re.findall(pattern_dd_mm_yyyy, year_str)
        match_yyyy_mm_dd = re.findall(pattern_yyyy_mm_dd, year_str)

        # Извлечение годов из найденных дат
        service_date_number_no_one.extend([date_str[-4:] for date_str in match_dd_mm_yyyy])  # Годы из формата "DD.MM.YYYY"
        service_date_number_no_one.extend([date_str[:4] for date_str in match_yyyy_mm_dd])    # Годы из формата "YYYY-MM-DD"

    # Оставляем только уникальные годы и преобразуем в целые числа
    service_date_number_no_one = list(set(int(year) for year in service_date_number_no_one if year.isdigit()))

    # Сортируем годы в порядке возрастания
    service_date_number_no_one.sort()

    # Если были пустые строки, добавляем None
    if empty_found_date_number_no_one:
        service_date_number_no_one.insert(0, None)

    # Преобразуем все значения в строки
    service_date_number_no_one = [str(year) for year in service_date_number_no_one]

    # service_years = db.session.query(db.func.year(Service.year)).distinct().all()
    # service_years = [str(year[0]) for year in service_years]

    """----------------------------------"""

    query = Service.query

    from sqlalchemy import not_, or_

    if year == 'No':
        year = None

    if date_number_no_one == 'No':
        date_number_no_one = None

    if year == 'None' and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd), Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))))
    elif year == 'None' and date_number_no_one:  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd))), Service.date_number_no_one.like(f'%{date_number_no_one}%'))
    elif year == 'None' and not date_number_no_one:  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd))))
    elif year and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))), Service.year.like(f'%{year}%'))
    elif not year and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))))
    elif year and date_number_no_one:
        # query = query.filter(db.func.year(Service.year) == year)
        query = query.filter(Service.year.like(f'%{year}%') | Service.date_number_no_one.like(f'%{date_number_no_one}%'))
    elif year and not date_number_no_one:
        # query = query.filter(db.func.year(Service.year) == year)
        query = query.filter(Service.year.like(f'%{year}%'))
    elif not year and date_number_no_one:
        # query = query.filter(db.func.year(Service.year) == year)
        query = query.filter(Service.date_number_no_one.like(f'%{date_number_no_one}%'))

    if keyword_one:
        if selected_column_one and hasattr(Service, selected_column_one):
            column = getattr(Service, selected_column_one)
            query = query.filter(column.like(f'%{keyword_one}%'))
        else:
            columns = [c for c in Service.__table__.columns.keys()]
            filters = [getattr(Service, col).like(f'%{keyword_one}%') for col in columns]
            query = query.filter(db.or_(*filters))

    if keyword_two:
        if selected_column_two and hasattr(Service, selected_column_two):
            column = getattr(Service, selected_column_two)
            query = query.filter(column.like(f'%{keyword_two}%'))
        else:
            columns = [c for c in Service.__table__.columns.keys()]
            filters = [getattr(Service, col).like(f'%{keyword_two}%') for col in columns]
            query = query.filter(db.or_(*filters))

    """Сортировка в cast(Service.id_id, Integer).asc() идёт с преобразованием в Integer"""
    from sqlalchemy import cast, Integer
    query = query.order_by(cast(Service.id_id, Integer).asc(), Service.year.asc())

    if year == 'None' and date_number_no_one == 'None':
        # Получение данных, которые не соответствуют формату "DD.MM.YYYY" и "YYYY-MM-DD"
        costs = db.session.query(Service.cost).filter(
            not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd), Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_1 = sum(float(cost[0]) for cost in costs if cost[0].replace('.', '', 1).isdigit())

        certificates = db.session.query(Service.certificate).filter(
            not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd), Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_2 = sum(float(cert[0]) for cert in certificates if cert[0].replace('.', '', 1).isdigit())

        certificates_no = db.session.query(Service.certificate_no).filter(
            not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd), Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_3 = sum(float(cert_no[0]) for cert_no in certificates_no if cert_no[0].replace('.', '', 1).isdigit())
    elif year and date_number_no_one:
        costs = db.session.query(Service.cost).filter(Service.year.like(f'%{year}%') | Service.date_number_no_one.like(f'%{year}%')).all()
        total_cost_1 = sum(float(cost[0]) for cost in costs if cost[0].replace('.', '', 1).isdigit())

        certificates = db.session.query(Service.certificate).filter(Service.year.like(f'%{year}%') | Service.date_number_no_one.like(f'%{year}%')).all()
        total_cost_2 = sum(float(cert[0]) for cert in certificates if cert[0].replace('.', '', 1).isdigit())

        certificates_no = db.session.query(Service.certificate_no).filter(Service.year.like(f'%{year}%') | Service.date_number_no_one.like(f'%{year}%')).all()
        total_cost_3 = sum(float(cert_no[0]) for cert_no in certificates_no if cert_no[0].replace('.', '', 1).isdigit())
    elif year == 'None' and date_number_no_one != 'None':
        # Получение данных, которые не соответствуют формату "DD.MM.YYYY" и "YYYY-MM-DD"
        costs = db.session.query(Service.cost).filter(
            not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_1 = sum(float(cost[0]) for cost in costs if cost[0].replace('.', '', 1).isdigit())

        certificates = db.session.query(Service.certificate).filter(
            not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_2 = sum(float(cert[0]) for cert in certificates if cert[0].replace('.', '', 1).isdigit())

        certificates_no = db.session.query(Service.certificate_no).filter(
            not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_3 = sum(float(cert_no[0]) for cert_no in certificates_no if cert_no[0].replace('.', '', 1).isdigit())
    elif year and not date_number_no_one:
        costs = db.session.query(Service.cost).filter(Service.year.like(f'%{year}%')).all()
        total_cost_1 = sum(float(cost[0]) for cost in costs if cost[0].replace('.', '', 1).isdigit())

        certificates = db.session.query(Service.certificate).filter(Service.year.like(f'%{year}%')).all()
        total_cost_2 = sum(float(cert[0]) for cert in certificates if cert[0].replace('.', '', 1).isdigit())

        certificates_no = db.session.query(Service.certificate_no).filter(Service.year.like(f'%{year}%')).all()
        total_cost_3 = sum(float(cert_no[0]) for cert_no in certificates_no if cert_no[0].replace('.', '', 1).isdigit())
    elif year != 'None' and date_number_no_one == 'None':
        # Получение данных, которые не соответствуют формату "DD.MM.YYYY" и "YYYY-MM-DD"
        costs = db.session.query(Service.cost).filter(
            not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_1 = sum(float(cost[0]) for cost in costs if cost[0].replace('.', '', 1).isdigit())

        certificates = db.session.query(Service.certificate).filter(
            not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_2 = sum(float(cert[0]) for cert in certificates if cert[0].replace('.', '', 1).isdigit())

        certificates_no = db.session.query(Service.certificate_no).filter(
            not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd)))
        ).all()

        total_cost_3 = sum(float(cert_no[0]) for cert_no in certificates_no if cert_no[0].replace('.', '', 1).isdigit())
    elif not year and date_number_no_one:
        costs = db.session.query(Service.cost).filter(Service.date_number_no_one.like(f'%{date_number_no_one}%')).all()
        total_cost_1 = sum(float(cost[0]) for cost in costs if cost[0].replace('.', '', 1).isdigit())

        certificates = db.session.query(Service.certificate).filter(Service.date_number_no_one.like(f'%{date_number_no_one}%')).all()
        total_cost_2 = sum(float(cert[0]) for cert in certificates if cert[0].replace('.', '', 1).isdigit())

        certificates_no = db.session.query(Service.certificate_no).filter(Service.date_number_no_one.like(f'%{date_number_no_one}%')).all()
        total_cost_3 = sum(float(cert_no[0]) for cert_no in certificates_no if cert_no[0].replace('.', '', 1).isdigit())
    else:
        total_cost_1 = db.session.query(db.func.sum(Service.cost)).scalar() or 0
        total_cost_2 = db.session.query(db.func.sum(Service.certificate)).scalar() or 0
        total_cost_3 = db.session.query(db.func.sum(Service.certificate_no)).scalar() or 0

    try:
        page = int(page)
        if not page > 0:
            page = 1
    except:
        page = 1

    offset = (page - 1) * per_page
    services = query.offset(offset).limit(per_page).all()
    total_services = query.count()
    total_pages = (total_services + per_page - 1) // per_page

    max_page_buttons = 5
    start_page = max(1, page - max_page_buttons // 2)
    end_page = min(total_pages, page + max_page_buttons // 2)

    if end_page - start_page < max_page_buttons - 1:
        if start_page > 1:
            end_page = min(total_pages, end_page + (max_page_buttons - (end_page - start_page)))
        else:
            start_page = max(1, end_page - (max_page_buttons - (end_page - start_page)))

    # return redirect(url_for('index'))
    return render_template(
        'index.html',
        services=services,
        total_cost_1=total_cost_1,
        total_cost_2=total_cost_2,
        total_cost_3=total_cost_3,
        selected_year=year,
        selected_date_number_no_one=date_number_no_one,
        selected_column_one=selected_column_one,
        selected_column_two=selected_column_two,
        keyword_one=keyword_one,
        keyword_two=keyword_two,
        page=page,
        total_pages=total_pages,
        start_page=start_page,
        end_page=end_page,
        service_years=service_years,
        service_date_number_no_one=service_date_number_no_one,
        form=csrf_token
    )

@app.route('/')
@login_required
def index():

    if not current_user.username in mas:
        flash('Неверное имя пользователя или пароль!', 'danger')
        return render_template('login.html', form=csrf_token)
    else:
        csrf_token = mas[current_user.username]

    total_pages_full = request.args.get('total_pages_full', None)

    if total_pages_full:
        per_page = 20
        query = Service.query
        total_services_full = query.count()
        page = (total_services_full + per_page - 1) // per_page
    else:
        page = request.args.get('page', 1, type=int)

    date_number_no_one = request.args.get('date_number_no_one', None)
    year = request.args.get('year', None)
    keyword_one = request.args.get('keyword_one', None)
    keyword_two = request.args.get('keyword_two', None)
    selected_column_one=request.args.get('selected_column_one', None)
    selected_column_two=request.args.get('selected_column_two', None)

    return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)

@app.route('/edit/<int:id>', methods=['GET'])
@login_required
def edit(id):
    if not current_user.username in mas:
        flash('Неверное имя пользователя или пароль!', 'danger')
        return render_template('login.html', form=csrf_token)
    else:
        csrf_token = mas[current_user.username]

    page = request.args.get('page', 1, type=int)
    keyword_one = request.args.get('keyword_one', None)
    keyword_two = request.args.get('keyword_two', None)
    selected_column_one=request.args.get('selected_column_one', None)
    selected_column_two=request.args.get('selected_column_two', None)
    selected_year = request.args.get('selected_year', 'No')
    selected_date_number_no_one = request.args.get('selected_date_number_no_one', 'No')

    # print('EDIT')
    # print(page)
    # print(keyword)
    # print(selected_column)
    # print(selected_year)

    service = Service.query.get_or_404(id)
    return render_template('edit.html', service=service, page=page, keyword_one=keyword_one, keyword_two=keyword_two, selected_column_one=selected_column_one, selected_column_two=selected_column_two, selected_year=selected_year, selected_date_number_no_one=selected_date_number_no_one, form=csrf_token)

@app.route('/add_edit', methods=['GET'])
@login_required
def add_edit():
    if not current_user.username in mas:
        flash('Неверное имя пользователя или пароль!', 'danger')
        return render_template('login.html', form=csrf_token)
    else:
        csrf_token = mas[current_user.username]

    page = request.args.get('page', 1, type=int)
    keyword_one = request.args.get('keyword_one', None)
    keyword_two = request.args.get('keyword_two', None)
    selected_column_one=request.args.get('selected_column_one', None)
    selected_column_two=request.args.get('selected_column_two', None)
    selected_year = request.args.get('selected_year', 'No')
    selected_date_number_no_one = request.args.get('selected_date_number_no_one', 'No')
    total_pages = request.args.get('total_pages', 1, type=int)

    # print('ADD_EDIT')
    # print(page)
    # print(keyword)
    # print(selected_column)
    # print(selected_year)

    return render_template('add.html', page=page, keyword_one=keyword_one, keyword_two=keyword_two, selected_column_one=selected_column_one, selected_column_two=selected_column_two, selected_year=selected_year, selected_date_number_no_one=selected_date_number_no_one, total_pages=total_pages, form=csrf_token)

@app.route('/edit/<int:id>', methods=['POST'])
@login_required
def update(id):
    if not current_user.username in mas:
        flash('Неверное имя пользователя или пароль!', 'danger')
        return render_template('login.html', form=csrf_token)
    else:
        csrf_token = mas[current_user.username]

    service = Service.query.get_or_404(id)
    service.id_id = request.form['id_id']
    service.name = request.form['name']
    service.snils = request.form['snils']
    service.location = request.form['location']
    service.address_p = request.form['address_p']
    service.address = request.form['address']
    service.benefit = request.form['benefit']
    service.number = request.form['number']

    service.year = request.form['year']
    try:
        if service.year:
            day = int(str(service.year).split('.')[0])
            month = int(str(service.year).split('.')[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                flash('Вы ввели неверный формат Даты выдачи сертификата. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
                # return redirect(url_for('index'))
                """---------------------------------"""

                page = request.args.get('page', 1, type=int)
                keyword_one = request.args.get('keyword_one', None)
                keyword_two = request.args.get('keyword_two', None)
                selected_column_one=request.args.get('selected_column_one', None)
                selected_column_two=request.args.get('selected_column_two', None)
                year = request.args.get('year', "")
                date_number_no_one = request.args.get('date_number_no_one', "")

                return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)
    except ValueError:
        flash('Вы ввели неверный формат Даты выдачи сертификата. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
        # return redirect(url_for('index'))
        """---------------------------------"""

        page = request.args.get('page', 1, type=int)
        keyword_one = request.args.get('keyword_one', None)
        keyword_two = request.args.get('keyword_two', None)
        selected_column_one=request.args.get('selected_column_one', None)
        selected_column_two=request.args.get('selected_column_two', None)
        year = request.args.get('year', "")
        date_number_no_one = request.args.get('date_number_no_one', "")

        return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)

    service.cost = request.form['cost']
    service.certificate = request.form['certificate']
    service.date_number_get = request.form['date_number_get']
    service.date_number_cancellation = request.form['date_number_cancellation']
    """ДОДЕЛАТЬ НОВЫЕ ГОДА"""
    service.date_number_no_one = request.form['date_number_no_one']
    try:
        if service.date_number_no_one:
            day = int(str(service.date_number_no_one).split('.')[0])
            month = int(str(service.date_number_no_one).split('.')[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                flash('Вы ввели неверный формат Даты решения об отказе в выдаче. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
                # return redirect(url_for('index'))
                """---------------------------------"""

                page = request.args.get('page', 1, type=int)
                keyword_one = request.args.get('keyword_one', None)
                keyword_two = request.args.get('keyword_two', None)
                selected_column_one=request.args.get('selected_column_one', None)
                selected_column_two=request.args.get('selected_column_two', None)
                year = request.args.get('year', "")
                date_number_no_one = request.args.get('date_number_no_one', "")

                return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)
    except ValueError:
        flash('Вы ввели неверный формат Даты решения об отказе в выдаче. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
        # return redirect(url_for('index'))
        """---------------------------------"""

        page = request.args.get('page', 1, type=int)
        keyword_one = request.args.get('keyword_one', None)
        keyword_two = request.args.get('keyword_two', None)
        selected_column_one=request.args.get('selected_column_one', None)
        selected_column_two=request.args.get('selected_column_two', None)
        year = request.args.get('year', "")
        date_number_no_one = request.args.get('date_number_no_one', "")

        return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)

    service.date_number_no_two = request.form['date_number_no_two']
    service.certificate_no = request.form['certificate_no']
    service.reason = request.form['reason']
    service.track = request.form['track']

    service.date_post = request.form['date_post']
    try:
        if service.date_post:
            day = int(str(service.date_post).split('.')[0])
            month = int(str(service.date_post).split('.')[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                flash('Вы ввели неверный формат Даты отправки почтой. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
                # return redirect(url_for('index'))
                """---------------------------------"""

                page = request.args.get('page', 1, type=int)
                keyword_one = request.args.get('keyword_one', None)
                keyword_two = request.args.get('keyword_two', None)
                selected_column_one=request.args.get('selected_column_one', None)
                selected_column_two=request.args.get('selected_column_two', None)
                year = request.args.get('year', "")
                date_number_no_one = request.args.get('date_number_no_one', "")

                return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)
    except ValueError:
        flash('Вы ввели неверный формат Даты отправки почтой. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
        # return redirect(url_for('index'))
        """---------------------------------"""

        page = request.args.get('page', 1, type=int)
        keyword_one = request.args.get('keyword_one', None)
        keyword_two = request.args.get('keyword_two', None)
        selected_column_one=request.args.get('selected_column_one', None)
        selected_column_two=request.args.get('selected_column_two', None)
        year = request.args.get('year', "")
        date_number_no_one = request.args.get('date_number_no_one', "")

        return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)

    service.comment = request.form['comment']
    service.color = request.form.get('color')

    if service.certificate == '0' and service.certificate_no == '0':
        service.color = '#dff0d8'

    db.session.commit()
    flash('Данные успешно обновлены!', 'success')
    # return redirect(url_for('index'))
    """---------------------------------"""

    page = request.args.get('page', 1, type=int)
    keyword_one = request.args.get('keyword_one', None)
    keyword_two = request.args.get('keyword_two', None)
    selected_column_one=request.args.get('selected_column_one', None)
    selected_column_two=request.args.get('selected_column_two', None)
    year = request.args.get('year', "")
    date_number_no_one = request.args.get('date_number_no_one', "")
    return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page, csrf_token)


@app.route('/update-color/<int:id>', methods=['POST'])
@login_required
def update_color(id):
    service = Service.query.get_or_404(id)
    data = request.json
    color = data.get('color')

    if color:
        service.color = color
        db.session.commit()
        return jsonify({
            'success': True,
            'id': service.id,
            'color': service.color
        })
    else:
        return jsonify({'success': False, 'message': 'No color provided'}), 400

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    service = Service.query.get_or_404(id)
    db.session.delete(service)
    db.session.commit()
    # flash('Данные успешно удалены!', 'success')
    return redirect(url_for('index'))

@app.route('/add', methods=['POST'])
@login_required
def add():
    if not current_user.username in mas:
        flash('Неверное имя пользователя или пароль!', 'danger')
        return render_template('login.html', form=csrf_token)
    else:
        csrf_token = mas[current_user.username]

    total_pages = request.args.get('total_pages', 1, type=int)

    # id_id = request.foыrm['id_id']
    name = request.form['name']
    snils = request.form['snils']
    location = request.form['location']
    address_p = request.form['address_p']
    address = request.form['address']
    benefit = request.form['benefit']
    number = request.form['number']

    year = request.form['year']
    try:
        if year:
            day = int(str(year).split('.')[0])
            month = int(str(year).split('.')[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                flash('Вы ввели неверный формат Даты выдачи сертификата. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
                # return redirect(url_for('index'))
                """---------------------------------"""

                # year = None
                # date_number_no_one = None
                # keyword_one = None
                # keyword_two = None
                # selected_column_one=None
                # selected_column_two=None
                # page = total_pages
                # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

                # Перенаправление после успешного добавления
                return redirect(url_for('index', page=total_pages, form=csrf_token))
    except ValueError:
        flash('Вы ввели неверный формат Даты выдачи сертификата. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
        # return redirect(url_for('index'))
        """---------------------------------"""

        # year = None
        # date_number_no_one = None
        # keyword_one = None
        # keyword_two = None
        # selected_column_one=None
        # selected_column_two=None
        # page = total_pages
        # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

        # Перенаправление после успешного добавления
        return redirect(url_for('index', page=total_pages, form=csrf_token))

    cost = request.form['cost']
    certificate = request.form['certificate']
    date_number_get = request.form['date_number_get']
    date_number_cancellation = request.form['date_number_cancellation']
    """ДОДЕЛАТЬ НОВЫЕ ГОДА"""
    date_number_no_one = request.form['date_number_no_one']
    try:
        if date_number_no_one:
            day = int(str(date_number_no_one).split('.')[0])
            month = int(str(date_number_no_one).split('.')[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                flash('Вы ввели неверный формат Даты решения об отказе в выдаче. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
                # return redirect(url_for('index'))
                """---------------------------------"""

                # year = None
                # date_number_no_one = None
                # keyword_one = None
                # keyword_two = None
                # selected_column_one=None
                # selected_column_two=None
                # page = total_pages
                # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

                # Перенаправление после успешного добавления
                return redirect(url_for('index', page=total_pages, form=csrf_token))
    except ValueError:
        flash('Вы ввели неверный формат Даты решения об отказе в выдаче. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
        # return redirect(url_for('index'))
        """---------------------------------"""

        # year = None
        # date_number_no_one = None
        # keyword_one = None
        # keyword_two = None
        # selected_column_one=None
        # selected_column_two=None
        # page = total_pages
        # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

        # Перенаправление после успешного добавления
        return redirect(url_for('index', page=total_pages, form=csrf_token))

    date_number_no_two = request.form['date_number_no_two']
    certificate_no = request.form['certificate_no']
    reason = request.form['reason']
    track = request.form['track']

    date_post = request.form['date_post']
    try:
        if date_post:
            day = int(str(date_post).split('.')[0])
            month = int(str(date_post).split('.')[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                flash('Вы ввели неверный формат Даты отправки почтой. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
                # return redirect(url_for('index'))
                """---------------------------------"""

                # year = None
                # date_number_no_one = None
                # keyword_one = None
                # keyword_two = None
                # selected_column_one=None
                # selected_column_two=None
                # page = total_pages
                # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

                # Перенаправление после успешного добавления
                return redirect(url_for('index', page=total_pages, form=csrf_token))
    except ValueError:
        flash('Вы ввели неверный формат Даты выдачи сертификата. Ожидаемый формат: ДД.ММ.ГГГГ.', 'danger')
        # return redirect(url_for('index'))
        """---------------------------------"""

        # year = None
        # date_number_no_one = None
        # keyword_one = None
        # keyword_two = None
        # selected_column_one=None
        # selected_column_two=None
        # page = total_pages
        # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

        # Перенаправление после успешного добавления
        return redirect(url_for('index', page=total_pages, form=csrf_token))

    comment = request.form['comment']
    color = request.form.get('color')

    if certificate == '0' and certificate_no == '0':
        color = '#dff0d8'

    from sqlalchemy import cast, Integer
    latest_service = Service.query.order_by(cast(Service.id_id, Integer).desc()).first()
    id_id = (int(latest_service.id_id) + 1) if latest_service else 1  # Преобразуем в int, если услуга найдена

    new_service = Service(id_id=id_id, name=name, snils=snils, location=location,
                        address_p=address_p, address=address, benefit=benefit,
                        number=number, year=year, cost=cost,
                        certificate=certificate, date_number_get=date_number_get,
                        date_number_cancellation=date_number_cancellation,
                        date_number_no_one=date_number_no_one, date_number_no_two=date_number_no_two, certificate_no=certificate_no,
                        reason=reason, track=track, date_post=date_post, comment=comment, color=color)
    db.session.add(new_service)
    db.session.commit()
    flash('Данные успешно добавлены!', 'success')


    """---------------------------------"""

    # year = None
    # date_number_no_one = None
    # keyword_one = None
    # keyword_two = None
    # selected_column_one=None
    # selected_column_two=None
    # page = total_pages
    # return skeleton(date_number_no_one, year, keyword_one, keyword_two, selected_column_one, selected_column_two, page)

    # Перенаправление после успешного добавления
    return redirect(url_for('index', page=total_pages, form=csrf_token))

# @app.route('/export-excel', methods=['GET'])
# @login_required
# def export_excel():
#     year = request.args.get('year', None)

#     query = Service.query

    # if year == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
    #     query = query.filter(Service.year.is_(None) | (Service.year == ''))
    # elif year:
    #     # query = query.filter(db.func.year(Service.year) == year)
    #     query = query.filter(Service.year.like(f'%{year}%'))

#     services = query.all()

#     df = pd.DataFrame([{
#         '№ п/п': service.id_id,
#         'ФИО заявителя': service.name,
#         'СНИЛС': service.snils,
#         'Район': service.location,
#         'Населённый пункт': service.address_p,
#         'Адрес': service.address,
#         'Льгота': service.benefit,
#         'Серия и номер': service.number,
#         'Дата выдачи сертификата': service.year,
#         'Размер выплаты': service.cost,
#         'Сертификат': service.certificate,
#         'Дата и номер решения о выдаче': service.date_number_get,
#         'Дата и № решения об аннулировании': service.date_number_cancellation,
#         'Дата и № решения об отказе в выдаче': service.date_number_no,
#         'Отказ в выдаче': service.certificate_no,
#         'Причина отказа': service.reason,
#         'ТРЕК': service.track,
#         'Дата отправки почтой': service.date_post,
#         'Комментарий': service.comment,
#         'Color': getattr(service, 'color', '')
#     } for service in services])

#     # Приводим колонки к числовому типу данных
#     df['Размер выплаты'] = pd.to_numeric(df['Размер выплаты'], errors='coerce')
#     df['Сертификат'] = pd.to_numeric(df['Сертификат'], errors='coerce')
#     df['Отказ в выдаче'] = pd.to_numeric(df['Отказ в выдаче'], errors='coerce')

#     # Расчет итогов
#     total_cost = df['Размер выплаты'].sum()
#     total_certificate = df['Сертификат'].sum()
#     total_certificate_no = df['Отказ в выдаче'].sum()

#     # Создание строки с итогами
#     totals_row = pd.DataFrame([{
#         '№ п/п': '',
#         'ФИО заявителя': '',
#         'СНИЛС': '',
#         'Район': '',
#         'Населённый пункт': '',
#         'Адрес': '',
#         'Льгота': '',
#         'Серия и номер': '',
#         'Дата выдачи сертификата': '',
#         'Размер выплаты': total_cost,
#         'Сертификат': total_certificate,
#         'Дата и номер решения о выдаче': '',
#         'Дата и № решения об аннулировании': '',
#         'Дата и № решения об отказе в выдаче': '',
#         'Отказ в выдаче': total_certificate_no,
#         'Причина отказа': '',
#         'ТРЕК': '',
#         'Дата отправки почтой': '',
#         'Комментарий': '',
#         'Color': ''
#     }])

#     # Добавление строки с итогами в DataFrame
#     df = pd.concat([df, totals_row], ignore_index=True)

#     # Создаем файл Excel
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Services')

#         # Настройка стиля
#         worksheet = writer.sheets['Services']

#         # Определяем стиль для границ
#         border_style = Border(left=Side(style='thin'),
#                             right=Side(style='thin'),
#                             top=Side(style='thin'),
#                             bottom=Side(style='thin'))

#         # Определяем стиль для заливки желтым цветом
#         yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

#         # Применяем границы ко всем ячейкам и цвет к ячейкам, где он задан
#         for row_num in range(2, worksheet.max_row + 1):  # Пропускаем заголовки
#             for col_num in range(1, worksheet.max_column + 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
#                 cell.border = border_style

#                 # Применяем цвет только к ячейкам данных
#                 color = df.iloc[row_num - 2]['Color']  # Сопоставление индексов DataFrame
#                 if color:
#                     cell.fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type="solid")

#         # Применяем границы к заголовкам
#         for col_num in range(1, worksheet.max_column + 1):
#             cell = worksheet.cell(row=1, column=col_num)
#             cell.border = border_style

#         # Применяем желтый цвет к строке с итогами
#         totals_row_num = worksheet.max_row
#         for col_num in range(1, worksheet.max_column + 1):
#             cell = worksheet.cell(row=totals_row_num, column=col_num)
#             cell.fill = yellow_fill
#             cell.border = border_style

#         # Удаляем столбец Color из Excel-файла
#         worksheet.delete_cols(df.columns.get_loc("Color") + 1)

#     output.seek(0)

#     response = send_file(output, as_attachment=True, download_name='services.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#     # Отправляем файл пользователю
#     return response

from celery import Celery
celery = Celery(app.name, broker='redis://172.18.11.104:6379', backend='redis://172.18.11.104:6379')

from flask_socketio import SocketIO
socketio = SocketIO(app)

@celery.task
def export_excel_task(sid, data):
    with app.app_context():  # Создание контекста приложения

        try:

            print(f"Начало экспорта для сессии {sid} с данными: {data}")

            year = data.get('year', None)

            date_number_no_one = data.get('date_number_no_one', None)

            query = Service.query

            # Регулярное выражение для формата "DD.MM.YYYY"
            pattern_dd_mm_yyyy = r'\b\d{2}\.\d{2}\.\d{4}\b'

            # Регулярное выражение для формата "YYYY-MM-DD"
            pattern_yyyy_mm_dd = r'\b\d{4}-\d{2}-\d{2}\b'

            from sqlalchemy import not_, or_

            # if year == 'None' and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
            #     # query = query.filter(Service.year.is_(None) | (Service.year == ''))
            #     query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd), Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))))
            #     print('POPAL_1')
            # elif year and year != 'No' and date_number_no_one and date_number_no_one != 'No':
            #     # query = query.filter(db.func.year(Service.year) == year)
            #     query = query.filter(Service.year.like(f'%{year}%') | Service.date_number_no_one.like(f'%{date_number_no_one}%'))
            #     print('POPAL_2')
            # elif year == 'No' and date_number_no_one == 'No':
            #     year = None
            #     date_number_no_one = None
            #     print('POPAL_3')
            # elif year == 'None' and date_number_no_one != 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
            #     # query = query.filter(Service.year.is_(None) | (Service.year == ''))
            #     query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd))), Service.date_number_no_one.like(f'%{date_number_no_one}%'))
            #     print('POPAL_4')
            # elif year and year != 'No' and not date_number_no_one and date_number_no_one == 'No':
            #     # query = query.filter(db.func.year(Service.year) == year)
            #     query = query.filter(Service.year.like(f'%{year}%'))
            #     print('POPAL_5')
            # elif year == 'No' and date_number_no_one != 'No':
            #     year = None
            #     query = query.filter(Service.date_number_no_one.like(f'%{date_number_no_one}%'))
            #     print('POPAL_6')
            # elif year != 'None' and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
            #     # query = query.filter(Service.year.is_(None) | (Service.year == ''))
            #     query = query.filter(not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))))
            #     print('POPAL_7')
            # elif year == 'No' and date_number_no_one and date_number_no_one != 'No':
            #     # query = query.filter(db.func.year(Service.year) == year)
            #     query = query.filter(Service.date_number_no_one.like(f'%{date_number_no_one}%'))
            #     print('POPAL_8')
            # elif year != 'No' and date_number_no_one == 'No':
            #     date_number_no_one = None
            #     query = query.filter(Service.year.like(f'%{year}%'))
            #     print('POPAL_9')

            if year == 'No':
                year = None

            if date_number_no_one == 'No':
                date_number_no_one = None

            if year == 'None' and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
                query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd), Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))))
            elif year == 'None' and date_number_no_one:  # Если year == 'None', фильтруем записи, у которых год == NULL
                query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd))), Service.date_number_no_one.like(f'%{date_number_no_one}%'))
            elif year == 'None' and not date_number_no_one:  # Если year == 'None', фильтруем записи, у которых год == NULL
                query = query.filter(not_(or_(Service.year.op('regexp')(pattern_dd_mm_yyyy), Service.year.op('regexp')(pattern_yyyy_mm_dd))))
            elif year and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
                query = query.filter(not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))), Service.year.like(f'%{year}%'))
            elif not year and date_number_no_one == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
                query = query.filter(not_(or_(Service.date_number_no_one.op('regexp')(pattern_dd_mm_yyyy), Service.date_number_no_one.op('regexp')(pattern_yyyy_mm_dd))))
            elif year and date_number_no_one:
                # query = query.filter(db.func.year(Service.year) == year)
                query = query.filter(Service.year.like(f'%{year}%') | Service.date_number_no_one.like(f'%{date_number_no_one}%'))
            elif year and not date_number_no_one:
                # query = query.filter(db.func.year(Service.year) == year)
                query = query.filter(Service.year.like(f'%{year}%'))
            elif not year and date_number_no_one:
                # query = query.filter(db.func.year(Service.year) == year)
                query = query.filter(Service.date_number_no_one.like(f'%{date_number_no_one}%'))

            services = query.all()
            print(f"Получено {len(services)} услуг.")

            df = pd.DataFrame([{
                '№ п/п': service.id_id,
                'ФИО заявителя': service.name,
                'СНИЛС': service.snils,
                'Район': service.location,
                'Населённый пункт': service.address_p,
                'Адрес': service.address,
                'Льгота': service.benefit,
                'Серия и номер': service.number,
                'Дата выдачи сертификата': service.year,
                'Размер выплаты': service.cost,
                'Сертификат': service.certificate,
                'Дата и номер решения о выдаче': service.date_number_get,
                'Дата и № решения об аннулировании': service.date_number_cancellation,
                'Дата решения об отказе в выдаче': service.date_number_no_one,
                '№ решения об отказе в выдаче': service.date_number_no_two,
                'Отказ в выдаче': service.certificate_no,
                'Причина отказа': service.reason,
                'ТРЕК': service.track,
                'Дата отправки почтой': service.date_post,
                'Комментарий': service.comment,
                'Color': getattr(service, 'color', '')
            } for service in services])

            # Приводим колонки к числовому типу данных
            df['Размер выплаты'] = pd.to_numeric(df['Размер выплаты'], errors='coerce')
            df['Сертификат'] = pd.to_numeric(df['Сертификат'], errors='coerce')
            df['Отказ в выдаче'] = pd.to_numeric(df['Отказ в выдаче'], errors='coerce')

            # Расчет итогов
            total_cost = df['Размер выплаты'].sum()
            total_certificate = df['Сертификат'].sum()
            total_certificate_no = df['Отказ в выдаче'].sum()

            # Создание строки с итогами
            totals_row = pd.DataFrame([{
                '№ п/п': '',
                'ФИО заявителя': '',
                'СНИЛС': '',
                'Район': '',
                'Населённый пункт': '',
                'Адрес': '',
                'Льгота': '',
                'Серия и номер': '',
                'Дата выдачи сертификата': '',
                'Размер выплаты': total_cost,
                'Сертификат': total_certificate,
                'Дата и номер решения о выдаче': '',
                'Дата и № решения об аннулировании': '',
                'Дата решения об отказе в выдаче': '',
                '№ решения об отказе в выдаче': '',
                'Отказ в выдаче': total_certificate_no,
                'Причина отказа': '',
                'ТРЕК': '',
                'Дата отправки почтой': '',
                'Комментарий': '',
                'Color': ''
            }])

            # Добавление строки с итогами в DataFrame
            df = pd.concat([df, totals_row], ignore_index=True)

            # Создаем файл Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Services')

                # Настройка стиля
                worksheet = writer.sheets['Services']

                # Определяем стиль для границ
                border_style = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))

                # Определяем стиль для заливки желтым цветом
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                import re
                def is_valid_hex_color(color):
                    # Регулярное выражение для проверки формата цвета
                    pattern = r'^#[0-9A-Fa-f]{6}$'
                    return bool(re.match(pattern, color))

                # Применяем границы ко всем ячейкам и цвет к ячейкам, где он задан
                for row_num in range(2, worksheet.max_row + 1):  # Пропускаем заголовки
                    for col_num in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = border_style

                        # Применяем цвет только к ячейкам данных
                        color = df.iloc[row_num - 2]['Color']  # Сопоставление индексов DataFrame
                        if color and is_valid_hex_color(color):
                            cell.fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type="solid")

                # Применяем границы к заголовкам
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.border = border_style

                # Применяем желтый цвет к строке с итогами
                totals_row_num = worksheet.max_row
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=totals_row_num, column=col_num)
                    cell.fill = yellow_fill
                    cell.border = border_style

                # Удаляем столбец Color из Excel-файла
                worksheet.delete_cols(df.columns.get_loc("Color") + 1)

            output.seek(0)

            import datetime
            date = str(datetime.datetime.now().date())

            # Получаем текущую директорию проекта
            project_dir = os.path.dirname(os.path.abspath(__file__))

            # Строим путь к папке file внутри проекта
            file_path = os.path.join(project_dir, 'file', f'services_{sid}_{date}.xlsx')

            # Сохраняем файл на диск
            with open(file_path, 'wb') as f:
                f.write(output.read())

            # Возвращаем путь к файлу для отправки клиенту
            filename = f"services_{sid}_{date}.xlsx"
            file_url = f"/file/{filename}"

            # # Emit the success event with file URL
            # socketio.emit('export_success', {'file_url': file_url, 'filename': filename}, room=sid)

            return file_url, filename

            # # Преобразуем файл в base64, чтобы отправить его через WebSocket
            # file_data = base64.b64encode(output.read()).decode('utf-8')
            # # return file_data

            # print(f"Отправка данных клиенту: {file_data}, имя файла: services.xlsx")

            # return file_data

        except Exception as e:
            print(f"Ошибка в задаче для сессии {sid}: {e}")
            socketio.emit('export_error', {'message': str(e)}, room=request.sid)

# Обработчик события для экспорта данных
@socketio.on('export_excel')
def handle_export_excel(data):
    sid = request.sid  # Идентификатор соединения

    task = export_excel_task.delay(sid, data)

    # Отправляем клиенту task_id для отслеживания задачи
    socketio.emit('export_started', {'task_id': task.id}, room=sid)

from celery.result import AsyncResult

@socketio.on('check_task_status')
def handle_check_task_status(data):
    task_id = data.get('task_id')

    # Проверяем статус задачи по task_id
    task_result = AsyncResult(task_id, app=celery)

    if task_result.state == 'PENDING':
        socketio.emit('task_status_pending', {'status': 'pending'}, room=request.sid)
    elif task_result.state == 'SUCCESS':
        # socketio.emit('task_status', {'status': 'success'}, room=request.sid)
        # socketio.emit('export_success', {'file_data': task_result.result, 'filename': 'services.xlsx'}, room=request.sid)
        socketio.emit('export_success', {'file_url': task_result.result[0], 'filename': task_result.result[1]}, room=request.sid)
    elif task_result.state == 'FAILURE':
        socketio.emit('task_status_failure', {'status': 'failure', 'error': str(task_result.info)}, room=request.sid)

from flask import send_from_directory
# Маршрут для скачивания файла
@app.route('/file/<filename>')
def download_file(filename):
    # Получаем текущую директорию проекта
    project_dir = os.path.dirname(os.path.abspath(__file__))

    # Строим путь к папке file внутри проекта
    directory = os.path.join(project_dir, 'file')

    return send_from_directory(directory, filename)

@socketio.on('disconnect')
def handle_disconnect():
    sid = request.sid
    print(f"Client disconnected: {sid}")

# """Nginx"""
# from waitress import serve
# if __name__ == '__main__':
#     print('Flask для Nginx запущен!')
#     serve(app, threads=10, host='172.18.11.103', port=8000)

# """Standart"""
# if __name__ == '__main__':
#     print('Flask запущен')
#     app.run(host='0.0.0.0', port=5000, debug=True)

"""SocketIo"""
if __name__ == '__main__':
    print('Flask запущен')
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)